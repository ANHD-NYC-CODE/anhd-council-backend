import csv
import openpyxl
import json
import io
import re
import types
import pyproj
import xlrd
from zipfile import ZipFile
from pyproj import *
import copy
from django.conf import settings
from .bbl import bbl
from .utility import merge
from core.utils.csv_helpers import count_csv_rows
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException


import logging

logger = logging.getLogger('app')


invalid_header_chars = ["\n", "\r", ' ', '-', '#', '.',
                        "'", '"', '_', '/', '(', ')', ':', "&", "’", '?']
replace_header_chars = [('%', 'pct')]
invalid_header_names = ["class"]
starts_with_numbers = re.compile(r'^(\d+)(.*)$')
only_numbers = re.compile(r'^\d+$')


def foreign_key_formatting(rows):
    for row in rows:
        if 'bbl' in row:
            row['bbl'] = str(row['bbl']).strip()
        if 'bin' in row:
            row['bin'] = str(row['bin']).strip()
        if 'registrationid' in row:
            row['registrationid'] = int(row['registrationid'])
        if 'documentid' in row:
            row['documentid'] = str(row['documentid']).strip()
        yield row


def flip_numbers(header):
    """
    str -> str
    If the header starts with numbers, it places
    those numbers at the end of the string.

    Columns in SQL cannot start with a number. This function will change
    a header named '2017values' to 'values2017'
    """
    match = starts_with_numbers.match(header)
    if match:
        if only_numbers.match(header):
            raise "Column names cannot be composed of all numbers"
        else:
            return (match.group(2) + match.group(1))
    else:
        return header


def clean_headers(headers):
    """
    str -> [str]
    turns header line into a list and fixes some commmon
    issues with column names
    """
    s = headers.lower()

    for char in invalid_header_chars:
        s = s.replace(char, '')
    for old, new in replace_header_chars:
        s = s.replace(old, new)
    for name in invalid_header_names:
        s = re.sub(r'\b{}\b'.format(re.escape(name)), name + '_name', s, 1)
    char_fixed_headers = [flip_numbers(x) for x in s.split(',')]
    fixed_headers = []

    for col_index in range(len(char_fixed_headers)):
        col_name = char_fixed_headers[col_index]
        if col_name in char_fixed_headers[:col_index]:
            occurances = char_fixed_headers[:col_index].count(col_name)
            fixed_headers.append(col_name + '_{}'.format(occurances + 1))
        else:
            fixed_headers.append(col_name)
    return fixed_headers

def from_geojson(file_path, pk='CounDist'):
    if isinstance(file_path, str):
        f = open(file_path, mode='r', encoding='utf-8', errors='replace')
    else:
        raise ValueError("from_geojson accepts Strings or Generators")

    with f:
        js = json.load(f)
        rows = js['features']

        for row in rows:
            # Add property information to the DB record
            # for key, value in row['properties'].items():
            #     row[key] = value
            #
            new_row = {}
            new_row['id'] = row['properties'][pk]

            row['properties']['id'] = row['properties'][pk]
            new_row['data'] = json.dumps(row)

            yield new_row


def from_dict_list_to_gen(list_rows):
    """
     from raw dict list [{header: value},...] to cleaned gen
    """
    for row in list_rows:
        headers = clean_headers(','.join(row.keys()))
        values = list(row.values())
        yield dict((headers[i], values[i]) for i in range(0, len(headers)))


def convert_xls_to_xlsx_workbook(file_path):
    # https://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx
    # first open using xlrd
    book_xls = xlrd.open_workbook(file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0, len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col +
                                1).value = sheet_xls.cell_value(row, col)

    return book_xlsx


def from_xlsx_file_to_gen(file_path, worksheet, update, skip_rows=0):
    if 'xlsx' in file_path.split('.')[-1]:
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = convert_xls_to_xlsx_workbook(file_path)
    worksheet = wb[worksheet]
    rows = worksheet.rows
    for i in range(skip_rows):
        next(rows)

    headers = clean_headers(','.join([c.value for c in next(rows)]))
    while not headers[0]:  # keeps going down if headers are blank
        headers = clean_headers(','.join([c.value for c in next(rows)]))

    for row in rows:
        values = [c.value for c in row]
        yield dict(zip(headers, values))

    if update:
        if update.total_rows:
            update.total_rows = update.total_rows + len(list(worksheet.rows))
        else:
            update.total_rows = len(list(worksheet.rows))
        update.save()
        update.save()


def from_csv_file_to_gen(file_path_or_generator, update=None, cleaner=None):
    logger.debug("Converting csv file to gen...")
    """
    input: String | Generator
    outs: Generator

    reads firstline as the headers and converts input into a stream of dicts
    """
    c = None
    if isinstance(file_path_or_generator, types.GeneratorType):
        gen_list = list(file_path_or_generator)

        f = io.StringIO('\n'.join(gen_list))
    elif isinstance(file_path_or_generator, str):
        f = open(file_path_or_generator, mode='r',
                 encoding='utf-8', errors='replace')
        if update:
            c = open(file_path_or_generator, mode='r',
                     encoding='utf-8', errors='replace')
    else:
        raise ValueError("from_csv_file_to_gen accepts Strings or Generators")

    if update and c:
        if update.total_rows:
            update.total_rows = count_csv_rows(file_path_or_generator)
        else:
            update.total_rows = count_csv_rows(file_path_or_generator)
        update.save()

    with f:
        if cleaner:
            lines = cleaner([*f])

            headers = clean_headers(next(lines))
            reader = csv.DictReader(lines, fieldnames=headers)
        else:
            headers = clean_headers(f.readline())
            reader = csv.DictReader(f, fieldnames=headers)

        logger.debug('Clearing NULL bytes')

        count = 0
        for row in reader:
            count = count + 1
            if count % settings.BATCH_SIZE == 0:
                logger.debug('processed {} csv rows'.format(count))

            yield row


def with_bbl(table, borough='borough', block='block', lot='lot', allow_blank=False):
    for row in table:
        try:
            yield merge(row, {'bbl': bbl(row[borough], row[block], row[lot])})
        except Exception as e:
            if allow_blank:
                yield row
            else:
                print(e)
                raise Exception(e)


p4j = '+proj=lcc +lat_1=40.66666666666666 +lat_2=41.03333333333333 +lat_0=40.16666666666666 +lon_0=-74 +x_0=300000 +y_0=0 +datum=NAD83 +units=us-ft +no_defs '
ny_state_plane = pyproj.Proj(p4j, preserve_units=True)
wgs84 = pyproj.Proj(init="epsg:4326", preserve_units=True)


def ny_state_coords_to_lat_lng(xcoord=0, ycoord=0):
    # returns (LNG, LAT)
    return pyproj.transform(ny_state_plane, wgs84, xcoord, ycoord)


def get_geo(row):
    try:
        if hasattr(row, '__iter__') and 'xcoord' in row and 'ycoord' in row:
            x = float(row['xcoord'])
            y = float(row['ycoord'])
        else:
            x = float(row.xcoord)
            y = float(row.ycoord)
        return ny_state_coords_to_lat_lng(xcoord=x, ycoord=y)
    except Exception as e:
        print('Geo error: {}'.format(e))
        logger.debug('Geo error: {}'.format(e))
        return (None, None)


def with_geo(table):
    logger.debug('processing geo...')
    count = 0
    for row in table:
        try:
            lng, lat = get_geo(row)
            count = count + 1
            if count % 10000 == 0:
                logger.debug('Processed {} geos'.format(count))
            yield merge(row, {'lng': lng, 'lat': lat})
        except Exception as e:
            logger.debug(e)
            yield merge(row, {'lng': None, 'lat': None})


def skip_fields(table, fields_to_skip):
    for row in table:
        for f in fields_to_skip:
            if f in row:
                del row[f]
        yield row
